from openpyxl import load_workbook


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


class ExcelStatusWorkbook:
    def __init__(self, path: str, status_column: str = "status"):
        self.path = path
        self.status_column = status_column
        self.workbook = load_workbook(path)
        self.sheet = self.workbook.active
        self.headers = [cell.value for cell in self.sheet[1]]
        self._ensure_status_column()
        self.summary = {
            "empty_rows": 0,
            "skipped_success_rows": [],
            "total_data_rows": self.sheet.max_row - 1,
        }

    def _ensure_status_column(self) -> None:
        if self.status_column not in self.headers:
            self.sheet.cell(row=1, column=len(self.headers) + 1, value=self.status_column)
            self.headers.append(self.status_column)

    def pending_rows(self) -> list[dict]:
        rows = []
        self.summary["empty_rows"] = 0
        self.summary["skipped_success_rows"] = []

        for row_number, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                self.summary["empty_rows"] += 1
                continue

            data = dict(zip(self.headers, row))
            if str(data.get(self.status_column)).strip().upper() == "SUCCESS":
                self.summary["skipped_success_rows"].append(row_number)
                continue

            rows.append({"row_number": row_number, "data": data})

        return rows

    def update_status(self, row_number: int, status: str) -> None:
        column_index = self.headers.index(self.status_column) + 1
        self.sheet.cell(row=row_number, column=column_index, value=status)
        self.workbook.save(self.path)
