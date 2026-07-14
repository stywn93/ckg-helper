import random
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook

PROVINCES = {
    "11": ("Aceh", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "71", "72", "73", "74", "75"]),
    "12": ("Sumatera Utara", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "71", "72", "73", "74", "75", "76", "77", "78"]),
    "13": ("Sumatera Barat", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "71", "72", "73", "74", "75"]),
    "14": ("Riau", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "71", "72", "73", "74"]),
    "15": ("Jambi", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "71", "72"]),
    "16": ("Sumatera Selatan", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "71", "72", "73"]),
    "17": ("Bengkulu", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "71", "72"]),
    "18": ("Lampung", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "71", "72"]),
    "19": ("Kep. Bangka Belitung", ["01", "02", "03", "04", "05", "06", "71"]),
    "21": ("Kepulauan Riau", ["01", "02", "03", "04", "05", "71", "72"]),
    "31": ("DKI Jakarta", ["01", "71", "72", "73", "74", "75"]),
    "32": ("Jawa Barat", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "71", "72", "73", "74", "75", "76", "77", "78"]),
    "33": ("Jawa Tengah", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "71", "72", "73", "74", "75"]),
    "34": ("DI Yogyakarta", ["01", "02", "03", "04", "71"]),
    "35": ("Jawa Timur", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "29", "71", "72", "73", "74", "75", "76", "77"]),
    "36": ("Banten", ["01", "02", "03", "04", "71", "72", "73"]),
    "51": ("Bali", ["01", "02", "03", "04", "05", "06", "07", "08", "71"]),
    "52": ("Nusa Tenggara Barat", ["01", "02", "03", "04", "05", "06", "07", "08", "71", "72"]),
    "53": ("Nusa Tenggara Timur", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "71", "72"]),
    "61": ("Kalimantan Barat", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "71", "72"]),
    "62": ("Kalimantan Tengah", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "71", "72"]),
    "63": ("Kalimantan Selatan", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "71", "72"]),
    "64": ("Kalimantan Timur", ["01", "02", "03", "04", "05", "06", "07", "71", "72", "73"]),
    "65": ("Kalimantan Utara", ["01", "02", "03", "04", "05", "71"]),
    "71": ("Sulawesi Utara", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "71", "72", "73", "74"]),
    "72": ("Sulawesi Tengah", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "71", "72"]),
    "73": ("Sulawesi Selatan", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "71", "72", "73", "74"]),
    "74": ("Sulawesi Tenggara", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "71", "72", "73", "74"]),
    "75": ("Gorontalo", ["01", "02", "03", "04", "05", "71"]),
    "76": ("Sulawesi Barat", ["01", "02", "03", "04", "05", "06"]),
    "81": ("Maluku", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "71", "72"]),
    "82": ("Maluku Utara", ["01", "02", "03", "04", "05", "06", "07", "08", "71", "72"]),
    "91": ("Papua", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", "25", "26", "27", "28", "71", "72"]),
    "92": ("Papua Barat", ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "71", "72"]),
    "93": ("Papua Selatan", ["01", "02", "03", "04"]),
    "94": ("Papua Tengah", ["01", "02", "03", "04", "05", "06", "07", "08"]),
    "95": ("Papua Pegunungan", ["01", "02", "03", "04", "05", "06", "07", "08"]),
    "96": ("Papua Barat Daya", ["01", "02", "03"]),
}

MALE_FIRST_NAMES = [
    "Adi", "Agus", "Ahmad", "Amir", "Andi", "Anton", "Ari", "Arif",
    "Bambang", "Bayu", "Budi", "Chandra", "Cipto", "Dani", "Dedi", "Deni",
    "Dimas", "Dwi", "Edi", "Eko", "Fajar", "Farid", "Gilang", "Guntur",
    "Gunawan", "Hadi", "Hendra", "Heru", "Imam", "Indra", "Irfan", "Joko",
    "Junaidi", "Kurniawan", "Kusnadi", "Lukman", "Luthfi", "Maman",
    "Maulana", "Mulyadi", "Nanda", "Nasrul", "Nurdin", "Oka", "Prayitno",
    "Purnomo", "Rahmat", "Ridwan", "Rudi", "Saiful", "Samsul", "Slamet",
    "Sugeng", "Suharto", "Sukri", "Syamsul", "Taufik", "Teguh", "Tri",
    "Untung", "Wahyu", "Wibowo", "Yudi", "Yusuf", "Zainal", "Zulkifli",
]

FEMALE_FIRST_NAMES = [
    "Aminah", "Ani", "Cahaya", "Dewi", "Dian", "Dini", "Dita", "Eka",
    "Elok", "Endang", "Eni", "Erna", "Farida", "Fatimah", "Fauziah", "Fitri",
    "Halimah", "Hesti", "Ida", "Ika", "Indah", "Irma", "Jamilah", "Kartika",
    "Kasih", "Kurnia", "Laili", "Lestari", "Lia", "Mainunah", "Marni",
    "Maya", "Nadia", "Ningsih", "Nirmala", "Nurul", "Puspita", "Putri",
    "Rahayu", "Rani", "Ratna", "Rina", "Rini", "Riska", "Safitri", "Sari",
    "Sartika", "Siska", "Siti", "Sri", "Suci", "Sumiati", "Susi",
    "Wati", "Wulandari", "Yanti", "Yuliana", "Yuliani", "Yuni", "Zainab",
    "Zulaikha",
]

LAST_NAMES = [
    "Ardianto", "Gunawan", "Handayani", "Hartono", "Harahap", "Hasibuan",
    "Hidayat", "Hutapea", "Irawan", "Kusuma", "Lubis", "Mahendra",
    "Manurung", "Marbun", "Nasution", "Nazara", "Nugraha", "Panggabean",
    "Permana", "Prakasa", "Pranoto", "Prasetyo", "Prayoga", "Purnama",
    "Santoso", "Saputra", "Setiawan", "Siahaan", "Simanjuntak", "Sinaga",
    "Siregar", "Situmorang", "Sudirman", "Susanto", "Telaumbanua",
    "Wibowo", "Wijaya", "Winata", "Zalukhu",
]


def generate_nik(gender: str, dob: date, province_code: str, regency_code: str, district_code: str, seq: int) -> str:
    day = dob.day
    if gender == "Perempuan":
        day += 40
    region_part = f"{province_code}{regency_code}{district_code}"
    date_part = f"{day:02d}{dob.month:02d}{dob.year % 100:02d}"
    seq_part = f"{seq:04d}"
    return f"{region_part}{date_part}{seq_part}"


def generate_name(gender: str) -> str:
    first_pool = MALE_FIRST_NAMES if gender == "Laki-laki" else FEMALE_FIRST_NAMES
    first = random.choice(first_pool)
    r = random.random()
    if r < 0.35:
        return first
    last = random.choice(LAST_NAMES)
    if r < 0.70:
        return f"{first} {last}"
    middle = random.choice(first_pool)
    return f"{first} {middle} {last}"


def generate_dob(min_age: int, max_age: int) -> date:
    today = date.today()
    year = random.randint(today.year - max_age, today.year - min_age)
    month = random.randint(1, 12)
    if month == 12:
        max_day = 31
    else:
        max_day = (date(year, month + 1, 1) - timedelta(days=1)).day
    day = random.randint(1, max_day)
    return date(year, month, day)


def generate_batch(
    count: int,
    male_ratio: float = 50.0,
    province_code: Optional[str] = None,
    regency_code: Optional[str] = None,
    min_age: int = 17,
    max_age: int = 65,
) -> list[dict]:
    persons = []
    used_niks = set()

    for _ in range(count):
        gender = "Laki-laki" if random.random() * 100 < male_ratio else "Perempuan"
        dob = generate_dob(min_age, max_age)
        name = generate_name(gender)

        if province_code and province_code in PROVINCES:
            prov = province_code
        else:
            prov = random.choice(list(PROVINCES.keys()))
        if regency_code and regency_code in PROVINCES[prov][1]:
            regency = regency_code
        else:
            regency = random.choice(PROVINCES[prov][1])
        district = f"{random.randint(1, 99):02d}"

        seq = random.randint(1, 9999)
        nik = generate_nik(gender, dob, prov, regency, district, seq)
        while nik in used_niks:
            seq = random.randint(1, 9999)
            nik = generate_nik(gender, dob, prov, regency, district, seq)
        used_niks.add(nik)

        persons.append({
            "nik": nik,
            "nama_lengkap": name,
            "tgl_lahir": dob.strftime("%Y-%m-%d"),
            "gender": gender,
        })

    return persons


def append_to_excel(excel_path: Path, rows: list[dict]) -> int:
    fields = ["nik", "nama_lengkap", "tgl_lahir", "gender"]

    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb.active
        existing = [cell.value for cell in ws[1]]
        col_map = {}
        for f in fields:
            if f in existing:
                col_map[f] = existing.index(f) + 1
            else:
                col_map[f] = len(existing) + 1
                existing.append(f)
                ws.cell(row=1, column=col_map[f], value=f)
    else:
        wb = Workbook()
        ws = wb.active
        col_map = {f: i + 1 for i, f in enumerate(fields)}
        ws.append(fields)

    for row in rows:
        ws.append([None] * len(ws[1]))
        r = ws.max_row
        for f in fields:
            ws.cell(row=r, column=col_map[f], value=row[f])

    wb.save(excel_path)
    return len(rows)


def list_provinces() -> str:
    lines = []
    for code, (name, _) in sorted(PROVINCES.items()):
        lines.append(f"  {code} = {name}")
    return "\n".join(lines)


def list_regencies(province_code: str) -> str:
    if province_code not in PROVINCES:
        return ""
    name, regencies = PROVINCES[province_code]
    lines = [f"Kode kota/kabupaten untuk {province_code} {name}:"]
    for code in regencies:
        lines.append(f"  {code}")
    return "\n".join(lines)
