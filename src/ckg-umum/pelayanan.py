import os
import traceback
import re

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright
from playwright_window_layout import launch_chromium_with_layout, pause_with_inspector_layout

load_dotenv()

STATUS_COLUMN = "status"
USERNAME_ENV = "CKG_USERNAME"
PASSWORD_ENV = "CKG_PASSWORD"
DEBUG_RAISE_ERRORS_ENV = "CKG_DEBUG_RAISE_ERRORS"
EXAMINATION_STATUS_OPTIONS = {
    "1": "Belum Pemeriksaan",
    "2": "Sedang Pemeriksaan",
    "3": "Selesai Pemeriksaan",
}
MONTH_LABELS = {
    "01": "Jan",
    "02": "Feb",
    "03": "Mar",
    "04": "Apr",
    "05": "Mei",
    "06": "Jun",
    "07": "Jul",
    "08": "Agt",
    "09": "Sep",
    "10": "Okt",
    "11": "Nov",
    "12": "Des",
}
MONTH_TO_NUMBER = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "Mei": 5,
    "Jun": 6,
    "Jul": 7,
    "Agt": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Des": 12,
}

def is_success_status(value) -> bool:
    return str(value).strip().upper() == "SUCCESS"


def get_required_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Environment variable {name} belum diisi.")
    return value


def load_rows_from_excel(path: str) -> tuple:
    workbook = load_workbook(path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if STATUS_COLUMN not in headers:
        sheet.cell(row=1, column=len(headers) + 1, value=STATUS_COLUMN)
        headers.append(STATUS_COLUMN)

    rows = []
    skipped_success_rows = []
    empty_rows = 0

    for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(row):
            empty_rows += 1
            continue
        row_data = dict(zip(headers, row))
        # skip pengecekan status dulu
        if is_success_status(row_data.get(STATUS_COLUMN)):
            skipped_success_rows.append(row_number)
            continue
        rows.append({"row_number": row_number, "data": row_data})

    summary = {
        "empty_rows": empty_rows,
        "skipped_success_rows": skipped_success_rows,
        "total_data_rows": sheet.max_row - 1,
    }

    return workbook, sheet, headers, rows, summary


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def update_row_status(
        workbook, sheet, headers: list, excel_path: str, row_number: int, status: str
) -> None:
    status_column_index = headers.index(STATUS_COLUMN) + 1
    sheet.cell(row=row_number, column=status_column_index, value=status)
    workbook.save(excel_path)


def prompt_examination_status() -> str:
    print("Pilih status pemeriksaan:")
    for option_number, option_label in EXAMINATION_STATUS_OPTIONS.items():
        print(f"{option_number}. {option_label}")

    while True:
        selected_option = input("Masukkan nomor pilihan (1-3): ").strip()
        if selected_option in EXAMINATION_STATUS_OPTIONS:
            return EXAMINATION_STATUS_OPTIONS[selected_option]
        print("Pilihan tidak valid. Masukkan nomor 1, 2, atau 3.")


def prepare_page(page) -> None:
    page.goto("https://sehatindonesiaku.kemkes.go.id/ckg-pelayanan")
    page.wait_for_load_state("networkidle")

    checkbox = page.locator("input[name='verify']")
    if checkbox.count() > 0:
        checkbox = page.locator("input[name='verify']")
        checkbox.set_checked(True, force=True)
        page.locator("button:has-text('Setuju')").click()
        page.wait_for_load_state("networkidle")

    sameLocation = page.locator("input[name='sameLocation']")
    if sameLocation.count() > 0:
        sameLocation = page.locator("input[name='sameLocation']")
        sameLocation.set_checked(True, force=True)
        page.locator("button:has-text('Simpan')").click()
        page.wait_for_load_state("networkidle")
    # print("end of prepare_page")


def search_patient(page, data: dict, row_number: int) -> None:
    prepare_page(page)
    # examination_status = prompt_examination_status()
    # page.locator("div.cursor-pointer.px-3").filter(has_text=examination_status).click()
    page.locator("div.cursor-pointer.px-3").filter(has_text="Sedang Pemeriksaan").click()
    page.locator("div.mx-input-wrapper").click()
    batas_awal = format_cell_value(data["batas_awal"])
    batas_akhir = format_cell_value(data["batas_akhir"])
    page.locator(f'td.cell[title="{batas_awal}"]').click()
    page.locator(f'td.cell[title="{batas_akhir}"]').click()

    page.locator("span:has-text('Nama')").click()
    page.get_by_text("Nama", exact=True).nth(0).click()
    page.locator("input#searchNik").fill(format_cell_value(data["nama"]))
    page.keyboard.press("Enter")
    page.wait_for_load_state("networkidle")
    page.locator("button:has-text('Mulai')").first.click()
    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(5000)
    for remaining_seconds in range(5, 0, -1):
        print(f"Menunggu halaman pemeriksaan tampil... {remaining_seconds} detik")
        page.wait_for_timeout(1000)
    # print("end of search_patient")


def do_demografi_dewasa(page, data: dict, row_number: int) -> None:
    print("Skrining Demografi Dewasa Dimulai")
    page.locator('[id="rowfrm000006"]').click()
    # soal 1
    page.locator("label").filter(
        has_text=format_cell_value(data["status_perkawinan"])
    ).click()
    # jika soal 1 jawabannya selain menikah, maka klik ini
    if data["status_perkawinan"] != "Menikah":
        page.locator("label").filter(
            has_text=format_cell_value(data["rencana_menikah"])
        ).click()

    # soal 2 atau 3
    page.locator("label").filter(
        has_text=format_cell_value(data["disabilitas"])
    ).click()
    page.locator("input:has-text('Kirim')").click()

    print("Skrining Demografi Dewasa Selesai")

def do_demografi_dewasa_perempuan(page, data: dict, row_number: int) -> None:
    print("Skrining Demografi Dewasa Perempuan Dimulai")
    page.locator('[id="rowfrm000007"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["status_perkawinan"])
    ).first.click()
    if data["status_perkawinan"] != "Menikah":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["rencana_menikah"])
        ).first.click()

    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["sedang_hamil"])
    ).first.click()

    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["disabilitas"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()

    print("Skrining Demografi Dewasa Perempuan Selesai")

def do_demografi_lansia(page, data: dict, row_number: int) -> None:
    print("Skrining Demografi Lansia Dimulai")
    page.locator('[id="rowfrm000007"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["status_perkawinan"])
    ).first.click()

    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["disabilitas"])
    ).first.click()

    page.locator("input:has-text('Kirim')").click()

    print("Skrining Demografi Lansia Selesai")


def do_risiko_kanker_usus(page, data: dict, row_number: int) -> None:
    print("Skrining Risiko Kanker Usus Dimulai")
    page.locator('[id="rowfrm000027"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["keluarga_kanker_usus"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["merokok"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Risiko Kanker Usus Selesai")


def do_risiko_tb(page, data: dict, row_number: int) -> None:
    print("Skrining Risiko TB Dimulai")
    page.locator('[id="rowfrm000180"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["batuk_tidak_sembuh"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Risiko Kanker Usus Selesai")


def do_hati(page, data: dict, row_number: int) -> None:
    print("Skrining Hati Dimulai")
    page.locator('[id="rowfrm000028"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["hati_hepatitis_b"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["ibu_hepatitis_b"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["seks_bukan_pasangan"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["transfusi_darah"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
        has_text=format_cell_value(data["hemodialisis"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pengguna_narkoba"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_106_ariaTitle'] label").filter(
        has_text=format_cell_value(data["odhiv"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_107_ariaTitle'] label").filter(
        has_text=format_cell_value(data["hati_hepatitis_c"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_108_ariaTitle'] label").filter(
        has_text=format_cell_value(data["kolesterol_tinggi"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Hati Selesai")

def do_leher_rahim(page, data: dict, row_number: int) -> None:
    print("Skrining Kanker Leher Rahim Dimulai")
    page.locator('[id="rowfrm000088"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pernah_seks"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kanker Leher Rahim Selesai")


def do_keswa(page, data: dict, row_number: int) -> None:
    print("Skrining Kesehatan Jiwa Dimulai")
    page.locator('[id="rowfrm000067"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tidak_bersemangat"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["merasa_tertekan"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["gugup_cemas"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["khawatir"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kesehatan Jiwa Selesai")


def do_risiko_kanker_paru(page, data: dict, row_number: int) -> None:
    print("Skrining Kanker Paru Dimulai")
    page.locator('[id="rowfrm000138"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["merokok_setahun_terakhir"])
    ).click()
    if data["merokok_setahun_terakhir"] == "Tidak":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["merokok_15_tahun"])
        ).click()
        page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
            has_text=format_cell_value(data["terpapar_rokok"])
        ).click()
        page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
            has_text=format_cell_value(data["kanker_paru_keluarga"])
        ).click()
        page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
            has_text=format_cell_value(data["gejala_batuk"])
        ).click()
        page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
            has_text=format_cell_value(data["tbc"])
        ).click()

    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["terpapar_rokok"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["kanker_paru_keluarga"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
        has_text=format_cell_value(data["gejala_batuk"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tbc"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kanker Paru Selesai")


def do_perilaku_merokok(page, data: dict, row_number: int) -> None:
    print("Skrining Perilaku Merokok Dimulai")
    page.locator('[id="rowfrm000064"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["merokok_setahun_terakhir_b"])
    ).click()

    if data["merokok_setahun_terakhir_b"] == "Ya":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["jenis_rokok"])
        ).click()
        page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
            format_cell_value(data["berapa_tahun"])
        )
        page.locator("input[aria-labelledby='sq_103_ariaTitle']").fill(
            format_cell_value(data["berapa_batang"])
        )

    elif data["merokok_setahun_terakhir_b"] == "Tidak":
        page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
            has_text=format_cell_value(data["pernah_merokok"])
        ).click()
        if data["pernah_merokok"] == "Ya":
            page.locator("input[aria-labelledby='sq_105_ariaTitle']").fill(
                format_cell_value(data["berapa_tahun_sebelumnya"])
            )
            page.locator("fieldset[aria-labelledby='sq_106_ariaTitle'] label").filter(
                has_text=format_cell_value(data["kapan_berhenti"])
            ).click()
    page.locator("fieldset[aria-labelledby='sq_107_ariaTitle'] label").filter(
        has_text=format_cell_value(data["terpapar_sebulan_terakhir"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Perilaku Merokok Selesai")


def do_aktivitas_fisik(page, data: dict, row_number: int) -> None:
    print("Skrining Aktivitas Fisik Dimulai")
    page.locator('[id="rowfrm000169"]').click()

    page.locator("div[aria-controls='sq_100i_list']").click()
    page.locator("#sq_100i_list [role='option']").filter(has_text=format_cell_value(data["aktivitas_domestik"])).click()
    if data["aktivitas_domestik"] == "Ya":
        page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
            format_cell_value(data["hari_domestik"])
        )
        page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
            format_cell_value(data["menit_domestik"])
        )

    page.locator("div[aria-controls='sq_103i_list']").click()
    page.locator("#sq_103i .sd-dropdown__value").click()
    page.locator("#sq_103i_list [role='option']").filter(
        has_text=format_cell_value(data["aktivitas_pekerjaan"])).click()
    if data["aktivitas_pekerjaan"] == "Ya":
        page.locator("input[aria-labelledby='sq_104_ariaTitle']").fill(
            format_cell_value(data["hari_pekerjaan"])
        )
        page.locator("input[aria-labelledby='sq_105_ariaTitle']").fill(
            format_cell_value(data["menit_pekerjaan"])
        )

    page.locator("div[aria-controls='sq_106i_list']").click()
    page.locator("#sq_106i .sd-dropdown__value").click()
    page.locator("#sq_106i_list [role='option']").filter(
        has_text=format_cell_value(data["aktivitas_perjalanan"])).click()
    if data["aktivitas_perjalanan"] == "Ya":
        page.locator("input[aria-labelledby='sq_107_ariaTitle']").fill(
            format_cell_value(data["hari_perjalanan"])
        )
        page.locator("input[aria-labelledby='sq_108_ariaTitle']").fill(
            format_cell_value(data["menit_perjalanan"])
        )

    page.locator("div[aria-controls='sq_109i_list']").click()
    page.locator("#sq_109i .sd-dropdown__value").click()
    page.locator("#sq_109i_list [role='option']").filter(
        has_text=format_cell_value(data["aktivitas_olahraga"])).click()
    if data["aktivitas_olahraga"] == "Ya":
        page.locator("input[aria-labelledby='sq_110_ariaTitle']").fill(
            format_cell_value(data["hari_olahraga"])
        )
        page.locator("input[aria-labelledby='sq_111_ariaTitle']").fill(
            format_cell_value(data["menit_olahraga"])
        )

    page.locator("div[aria-controls='sq_112i_list']").click()
    page.locator("#sq_112i .sd-dropdown__value").click()

    page.locator("#sq_112i_list [role='option']").filter(
        has_text=format_cell_value(data["aktivitas_kerja_berat"])).click()
    if data["aktivitas_kerja_berat"] == "Ya":
        page.locator("input[aria-labelledby='sq_113_ariaTitle']").fill(
            format_cell_value(data["hari_kerja_berat"])
        )
        page.locator("input[aria-labelledby='sq_114_ariaTitle']").fill(
            format_cell_value(data["menit_kerja_berat"])
        )

    page.locator("div[aria-controls='sq_115i_list']").click()
    page.locator("#sq_115i .sd-dropdown__value").click()
    page.locator("#sq_115i_list [role='option']").filter(
        has_text=format_cell_value(data["aktivitas_olahraga_berat"])).click()
    if data["aktivitas_olahraga_berat"] == "Ya":
        page.locator("input[aria-labelledby='sq_116_ariaTitle']").fill(
            format_cell_value(data["hari_olahraga_berat"])
        )
        page.locator("input[aria-labelledby='sq_117_ariaTitle']").fill(
            format_cell_value(data["menit_olahraga_berat"])
        )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Aktivitas Fisik Selesai")


# ---------- Pelayanan Oleh Nakes ------------

def do_pemeriksaan_check(page, selector: str, checked: bool) -> None:
    checkbox = page.locator(selector)
    if checked is False and checkbox.is_checked():
        # print("status pemeriksaan is False")
        checkbox.click(force=True)
        page.get_by_role("button", name="Tidak Periksa").click()
    elif checked is True:
        # print("status pemeriksaan is True")
        # print("selector ", selector)
        # page.pause()
        if checkbox.is_checked() != checked:
            checkbox.click()
        # checkbox.click()
        # page.locator(selector).set_checked(checked, force=True)

def do_gizi_laki(page, data: dict, row_number: int) -> None:
    print("Skrining Gizi Laki dimulai")
    page.locator('[id="rowfrm000093"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(format_cell_value(data["berat_badan"]))
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(format_cell_value(data["tinggi_badan"]))
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(format_cell_value(data["lingkar_perut"]))
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Gizi Laki selesai")

def do_gizi_perempuan(page, data: dict, row_number: int) -> None:
    print("Skrining Gizi Perempuan dimulai")
    page.locator('[id="rowfrm000051"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(format_cell_value(data["berat_badan"]))
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(format_cell_value(data["tinggi_badan"]))
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(format_cell_value(data["lingkar_perut"]))
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Gizi Perempuan selesai")

def do_skilas_penurunan_kognitif(page, data: dict, row_number: int) -> None:
    print("Skrining SKILAS Penurunan Kognitif dimulai")
    page.locator('[id="rowfrm000029"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["mengingat_3_kata"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tanggal_berapa_dimana"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["ulangi_3_kata_sebelumnya"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining SKILAS Penurunan Kognitif selesai")

def do_skilas_mobilisasi(page, data: dict, row_number: int) -> None:
    print("Skrining SKILAS Mobilisasi dimulai")
    page.locator('[id="rowfrm000032"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["berdiri_di_kursi"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining SKILAS Mobilisasi selesai")

def do_skilas_malnutrisi(page, data: dict, row_number: int) -> None:
    print("Skrining SKILAS Mobilisasi dimulai")
    page.locator('[id="rowfrm000034"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["berat_badan_berkurang"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["hilang_nafsu_makan"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["lila_kurang_21cm"])
    ).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining SKILAS Mobilisasi selesai")

def do_skilas_depresi(page, data: dict, row_number: int) -> None:
    print("Skrining SKILAS Gejala Depresi dimulai")
    page.locator('[id="rowfrm000038"]').click()
    page.locator("div[aria-controls='sq_100i_list']").click()
    page.locator("#sq_100i_list [role='option']").filter(
        has_text=format_cell_value(data["2_minggu_terakhir_sedih"])).click()
    page.locator("div[aria-controls='sq_101i_list']").click()
    page.locator("#sq_101i_list [role='option']").filter(
        has_text=format_cell_value(data["2_minggu_sedikit_minat"])).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining SKILAS Gejala Depresi selesai")

def do_gula_darah_dewasa(page, data: dict, row_number: int) -> None:
    print("Skrining Gula Darah Dewasa dimulai")
    do_pemeriksaan_check(page, "input#hasil-lab-0-1", True)
    page.locator('[id="rowfrm000256"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pernah_diabetes"])
    ).click()
    if(data["pernah_diabetes"] == "Ya"):
        page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
            format_cell_value(data["total_bulan_diabetes"])
        )
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
        format_cell_value(data["gula_darah_sewaktu"])
    )
    if (data["pernah_diabetes"] == "Tidak"):
        page.locator("input[aria-labelledby='sq_103_ariaTitle']").fill(
            format_cell_value(data["gula_darah_sewaktu_2"])
        )
    page.locator("input[aria-labelledby='sq_104_ariaTitle']").fill(
        format_cell_value(data["gula_darah_puasa"])
    )
    page.locator("input[aria-labelledby='sq_105_ariaTitle']").fill(
        format_cell_value(data["gula_darah_pp"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Gula Darah Dewasa selesai")

def do_tekanan_darah_dewasa(page, data: dict, row_number: int) -> None:
    print("Skrining Tekanan Darah Dewasa dimulai")
    do_pemeriksaan_check(page, "input#hasil-lab-0-2", True)
    page.locator('[id="rowfrm000265"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pernah_hipertensi"])
    ).click()
    if(data["pernah_hipertensi"] == "Ya"):
        page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
            format_cell_value(data["total_bulan_hipertensi"])
        )
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
        format_cell_value(data["sistolik"])
    )
    page.locator("input[aria-labelledby='sq_103_ariaTitle']").fill(
        format_cell_value(data["diastolik"])
    )
    page.locator("input[aria-labelledby='sq_104_ariaTitle']").fill(
        format_cell_value(data["sistolik_2"])
    )
    page.locator("input[aria-labelledby='sq_105_ariaTitle']").fill(
        format_cell_value(data["diastolik_2"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Tekanan Darah Dewasa selesai")

def do_risiko_tb(page, data: dict, row_number: int) -> None:
    print("Skrining Risiko TB dimulai")
    do_pemeriksaan_check(page, "input#hasil-lab-1-0", True)
    page.locator('[id="rowfrm000182"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pernah_batuk_tidak_sembuh"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["bb_turun"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["demam_hilang_timbul"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["berkeringat_malam"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pembesaran_getah_bening"])
    ).click()
    page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
        has_text=format_cell_value(data["radiografi_toraks"])
    ).click()
    if(data["radiografi_toraks"] == "Ya"):
        value = format_cell_value(data["hasil_rontgen"])
        page.locator(
            "fieldset[aria-labelledby='sq_106_ariaTitle'] label"
        ).filter(
            has_text=re.compile(rf"^{re.escape(value)}$")
        ).click()

    page.locator("input:has-text('Kirim')").click()
    print("Skrining Risiko TB selesai")
    # page.pause()

def do_tb(page, data: dict, row_number: int) -> None:
    print("Skrining TB dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-1-1']", True)
    page.locator('[id="rowfrm000184"]').click()

    page.locator("div[aria-controls='sq_100i_list']").click()
    page.locator("#sq_100i_list [role='option']").filter(has_text=format_cell_value(data["kontak_tbc"])).click()
    if data["kontak_tbc"] == "Riwayat kontak serumah" or data["kontak_tbc"] == "Riwayat kontak erat":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["jenis_tbc"])
        ).click()
    page.locator("div[aria-controls='sq_102i_list']").click()
    page.locator("#sq_102i_list [role='option']").filter(has_text=format_cell_value(data["metode_pemeriksaan_tbc"])).click()
    if data["metode_pemeriksaan_tbc"] == "TCM":
        # print("TCM")
        page.locator("div#sq_103i.sd-input.sd-dropdown").click()
        page.locator("#sq_103i_list [role='option']").filter(
            has_text=format_cell_value(data["hasil_pemeriksaan_tbc"])).click()
    elif data["metode_pemeriksaan_tbc"] == "BTA":
        page.locator("div[aria-controls='sq_104i_list']").click()
        page.locator("#sq_104i_list [role='option']").filter(
            has_text=format_cell_value(data["hasil_pemeriksaan_tbc"])).click()
    elif data["metode_pemeriksaan_tbc"] == "NPOC":
        page.locator("div[aria-controls='sq_105i_list']").click()
        page.locator("#sq_105i_list [role='option']").filter(
            has_text=format_cell_value(data["hasil_pemeriksaan_tbc"])).click()
    # page.pause()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining TB selesai")

def do_frambusia(page, data: dict, row_number: int) -> None:
    print("Skrining Frambusia dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-2-0']", True)
    page.locator('[id="rowfrm000199"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["ada_papul"])
    ).click()
    if data["ada_papul"] == "Suspek frambusia":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["hasil_pemeriksaan_rdt"])
        ).first.click()
    # page.pause()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Frambusia selesai")

def do_kusta(page, data: dict, row_number: int) -> None:
    print("Skrining Kusta dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-2-1']", True)
    page.locator('[id="rowfrm000198"]').click()
    page.locator("div[aria-controls='sq_100i_list']").click()
    page.locator("#sq_100i_list [role='option']").filter(has_text=format_cell_value(data["bercak_putih"])).click()
    if data["bercak_putih"] == "Meragukan":
        page.locator("div[aria-controls='sq_101i_list']").click()
        page.locator("#sq_101i_list [role='option']").filter(has_text=format_cell_value(data["hasil_bta"])).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kusta selesai")

def do_skabies(page, data: dict, row_number: int) -> None:
    print("Skrining Skabies dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-2-2']", True)
    page.locator('[id="rowfrm000201"]').click()
    page.locator("div[aria-controls='sq_100i_list']").click()
    page.locator("#sq_100i_list [role='option']").filter(has_text=format_cell_value(data["ada_ruam"])).click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Skabies selesai")

def do_telinga_mata(page, data: dict, row_number: int) -> None:
    print("Skrining Telinga dan Mata dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-3-0']", True)
    page.locator('[id="rowfrm000099"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["serumen_impaksi"])
    ).first.click()
    page.locator("div[aria-controls='sq_101i_list']").click()
    page.locator("#sq_101i_list [role='option']").filter(has_text=format_cell_value(data["infeksi_telinga"])).first.click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tajam_pendengaran"])
    ).first.click()
    if data["tajam_pendengaran"] == "Curiga gangguan pendengaran":
        page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
            has_text=format_cell_value(data["tes_penala"])
        ).first.click()
    page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tajam_penglihatan"])
    ).first.click()
    # print(data["tajam_penglihatan"])
    # page.pause()
    if data["tajam_penglihatan"] == "Curiga gangguan penglihatan (visus <6/12)":
        page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
            has_text=format_cell_value(data["hasil_visus"])
        ).first.click()
        if data["hasil_visus"] != "Normal (visus 6/6 - 6/12)":
            page.locator("fieldset[aria-labelledby='sq_106_ariaTitle'] label").filter(
                has_text=format_cell_value(data["pinhole"])
            ).first.click()
            if data["pinhole"] == "Visus membaik":
                page.locator("fieldset[aria-labelledby='sq_107_ariaTitle'] label").filter(
                    has_text=format_cell_value(data["hasil_refraksi"])
                ).first.click()
            else:
                page.locator("fieldset[aria-labelledby='sq_108_ariaTitle'] label").filter(
                    has_text=format_cell_value(data["funduskopi"])
                ).first.click()
    page.locator("fieldset[aria-labelledby='sq_109_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pupil"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Telinga dan Mata selesai")

def do_karies(page, data: dict, row_number: int) -> None:
    print("Skrining Karies dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-4-0']", True)
    page.locator('[id="rowfrm000055"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["gigi_karies"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["gigi_hilang"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Karies selesai")

def do_periodontal(page, data: dict, row_number: int) -> None:
    print("Skrining Periodontal dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-4-1']", True)
    page.locator('[id="rowfrm000056"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["penyakit_periodontal"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["gigi_goyang"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Periodontal selesai")

def do_ppok(page, data: dict, row_number: int) -> None:
    print("Skrining PPOK dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-5-0']", True)
    page.locator('[id="rowfrm000101"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["ppok_merokok"])
    ).first.click()
    if data["ppok_merokok"] == "Iya":
        page.locator("div[aria-controls='sq_101i_list']").click()
        page.locator("#sq_101i_list [role='option']").filter(
            has_text=format_cell_value(data["bungkus_per_tahun"])).first.click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["nafas_pendek"])).first.click()
    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["mempunyai_dahak"])).first.click()
    page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
        has_text=format_cell_value(data["batuk_tanpa_flu"])).first.click()
    page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
        has_text=format_cell_value(data["periksa_spirometri"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining PPOK selesai")

def do_kadar_co(page, data: dict, row_number: int) -> None:
    print("Skrining Kadar CO dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-6-0']", True)
    page.locator('[id="rowfrm000186"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["kadar_co"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kadar CO selesai")

def do_lipid(page, data: dict, row_number: int) -> None:
    print("Skrining Lipid dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-7-0']", True)
    page.locator('[id="rowfrm000047"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["kolesterol"])
    )
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
        format_cell_value(data["hdl"])
    )
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
        format_cell_value(data["ldl"])
    )
    page.locator("input[aria-labelledby='sq_103_ariaTitle']").fill(
        format_cell_value(data["trigliserida"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Lipid selesai")

def do_fibrosis(page, data: dict, row_number: int) -> None:
    print("Skrining Fibrosis dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-7-1']", True)
    page.locator('[id="rowfrm000045"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["sgot"])
    )
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
        format_cell_value(data["trombosit"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Fibrosis selesai")

def do_hepatitis(page, data: dict, row_number: int) -> None:
    print("Skrining Hepatitis dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-7-2']", True)
    page.locator('[id="rowfrm000044"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["hepatitis_b"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["hepatitis_c"])
    ).first.click()
    if data["hepatitis_c"] == "Anti HCV Reaktif":
        page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
            has_text=format_cell_value(data["vl_hepatitis_c"])
        ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Hepatitis selesai")

def do_fungsi_ginjal(page, data: dict, row_number: int) -> None:
    print("Skrining Fungsi Ginjal dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-7-3']", True)
    page.locator('[id="rowfrm000244"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["kreatinin"])
    )
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
        format_cell_value(data["ureum"])
    )
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
        format_cell_value(data["usia"])
    )
    page.locator("input[aria-labelledby='sq_103_ariaTitle']").fill(
        format_cell_value(data["e_lfg"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Fungsi Ginjal selesai")

def do_fungsi_ginjal_perempuan(page, data: dict, row_number: int) -> None:
    print("Skrining Fungsi Ginjal dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-7-3']", True)
    page.locator('[id="rowfrm000245"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["kreatinin"])
    )
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
        format_cell_value(data["ureum"])
    )
    page.locator("input[aria-labelledby='sq_102_ariaTitle']").fill(
        format_cell_value(data["usia"])
    )
    page.locator("input[aria-labelledby='sq_103_ariaTitle']").fill(
        format_cell_value(data["e_lfg"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Fungsi Ginjal selesai")

def do_kerusakan_ginjal(page, data: dict, row_number: int) -> None:
    print("Skrining Kerusakan Ginjal dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-7-4']", True)
    page.locator('[id="rowfrm000248"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["albumin"])
    )
    page.locator("input[aria-labelledby='sq_101_ariaTitle']").fill(
        format_cell_value(data["kreatinin_urin"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kerusakan Ginjal selesai")

def do_kanker_payudara(page, data: dict, row_number: int) -> None:
    print("Skrining Kanker Payudara dimulai")
    # do_pemeriksaan_check(page, "label[for='hasil-lab-8-0']", True)
    page.locator('[id="rowfrm000059"]').click()
    page.locator("div[aria-controls='sq_100i_list']").click()
    page.locator("#sq_100i_list [role='option']").filter(
        has_text=format_cell_value(data["pemeriksaan_payudara"])).first.click()
    if data["pemeriksaan_payudara"] == "SADANIS":
        page.locator("div[aria-controls='sq_101i_list']").click()
        page.locator("#sq_101i_list [role='option']").filter(
            has_text=format_cell_value(data["hasil_sadanis"])).first.click()
    else:
        page.locator("div[aria-controls='sq_102i_list']").click()
        page.locator("#sq_102i_list [role='option']").filter(
            has_text=format_cell_value(data["hasil_usg_payudara"])).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kanker Payudara selesai")

def do_hpv_dna(page, data: dict, row_number: int) -> None:
    print("Skrining HPV DNA dimulai")
    # do_pemeriksaan_check(page, "label[for='hasil-lab-8-0']", True)
    page.locator('[id="rowfrm000061"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pemeriksaan_hpv_dna"])
    ).first.click()
    if data["pemeriksaan_hpv_dna"] == "HPV Positif":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["hpv_16"])
        ).first.click()
        page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
            has_text=format_cell_value(data["hpv_18"])
        ).first.click()
        page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
            has_text=format_cell_value(data["hpv_52"])
        ).first.click()
        page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
            has_text=format_cell_value(data["hpv_enkogenik_lain"])
        ).first.click()

    page.locator("input:has-text('Kirim')").click()
    print("Skrining HPV DNA selesai")

def do_inspekulo_iva(page, data: dict, row_number: int) -> None:
    print("Skrining Inspekulo dan IVA dimulai")
    # do_pemeriksaan_check(page, "label[for='hasil-lab-8-0']", True)
    page.locator('[id="rowfrm000060"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["inspekulo"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["iva"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Inspekulo dan IVA selesai")

def do_jantung(page, data: dict, row_number: int) -> None:
    print("Skrining Jantung dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-8-0']", True) #butuh penyesuaian berdasarkan siklus hidupnya atau bisa juga tahapan ini dilewati dengan asumsi saat pertama kali dibuka semuanya akan Ya
    page.locator('[id="rowfrm000057"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["ekg"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["pemeriksaan_ekg"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Jantung selesai")

def do_kanker_usus(page, data: dict, row_number: int) -> None:
    print("Skrining Kanker Usus dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-9-0']", True)
    page.locator('[id="rowfrm000050"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["bersedia_colok_dubur"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["colok_dubur"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["darah_samar"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kanker Usus selesai")

def do_kanker_paru(page, data: dict, row_number: int) -> None:
    print("Skrining Kanker Paru dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-10-0']", True)
    page.locator('[id="rowfrm000041"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["kanker_paru"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
        has_text=format_cell_value(data["keluarga_kanker"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
        has_text=format_cell_value(data["riwayat_merokok"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tempat_kerja_karsinogenik"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_104_ariaTitle'] label").filter(
        has_text=format_cell_value(data["tempat_tinggal_potensi_tinggi"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_105_ariaTitle'] label").filter(
        has_text=format_cell_value(data["lingkungan_tidak_sehat"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_106_ariaTitle'] label").filter(
        has_text=format_cell_value(data["penyakit_paru_kronik"])
    ).first.click()
    page.locator("fieldset[aria-labelledby='sq_107_ariaTitle'] label").filter(
        has_text=format_cell_value(data["foto_torax"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Kanker Paru selesai")

def do_catin_perempuan(page, data: dict, row_number: int) -> None:
    print("Skrining Catin Perempuan dimulai")
    # do_pemeriksaan_check(page, "label[for='hasil-lab-10-0']", True)
    page.locator('[id="rowfrm000205"]').click()
    page.locator("input[aria-labelledby='sq_100_ariaTitle']").fill(
        format_cell_value(data["hemoglobin"])
    )
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Catin Perempuan selesai")

def do_hiv(page, data: dict, row_number: int) -> None:
    print("Skrining HIV dimulai")
    do_pemeriksaan_check(page, "label[for='hasil-lab-11-0']", True)
    page.locator('[id="rowfrm000188"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["rapid_test"])
    ).first.click()
    if data["rapid_test"] == "Reaktif":
        page.locator("fieldset[aria-labelledby='sq_101_ariaTitle'] label").filter(
            has_text=format_cell_value(data["r2_hiv"])
        ).first.click()
        if data["r2_hiv"] == "Reaktif":
            page.locator("fieldset[aria-labelledby='sq_103_ariaTitle'] label").filter(
                has_text=format_cell_value(data["r3_hiv"])
            ).first.click()
        elif data["r2_hiv"] == "Non Reaktif":
            page.locator("fieldset[aria-labelledby='sq_102_ariaTitle'] label").filter(
                has_text=format_cell_value(data["r1_hiv_ulang"])
            ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining HIV selesai")

def do_sifilis(page, data: dict, row_number: int) -> None:
    print("Skrining Sifilis dimulai")
    # do_pemeriksaan_check(page, "label[for='hasil-lab-11-1']", True)
    page.locator('[id="rowfrm000191"]').click()
    page.locator("fieldset[aria-labelledby='sq_100_ariaTitle'] label").filter(
        has_text=format_cell_value(data["rapid_test_sifilis"])
    ).first.click()
    page.locator("input:has-text('Kirim')").click()
    print("Skrining Sifilis selesai")

def main():
    excel_path = "dataset/pelayanan.xlsx"
    username = get_required_env(USERNAME_ENV)
    password = get_required_env(PASSWORD_ENV)
    # examination_status = prompt_examination_status()
    workbook, sheet, headers, data_rows, excel_summary = load_rows_from_excel(
        excel_path
    )
    if not data_rows:
        skipped_success_rows = excel_summary["skipped_success_rows"]
        if skipped_success_rows:
            print(
                "Tidak ada data yang perlu diproses. "
                f"Baris {skipped_success_rows} dilewati karena status sudah SUCCESS."
            )
            print("Kosongkan kolom status untuk memproses ulang baris tersebut.")
        else:
            print("Tidak ada data pada file Excel.")
            print("")
        return

    with sync_playwright() as p:
        browser, window_layout = launch_chromium_with_layout(p)
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        page.goto("https://sehatindonesiaku.kemkes.go.id/ckg-pelayanan")
        page.locator("input#email").fill(username)
        page.locator("input#password").fill(password)
        pause_with_inspector_layout(page, window_layout)

        failed_rows = []

        for row_entry in data_rows:
            index = row_entry["row_number"]
            data = row_entry["data"]
            try:
                search_patient(page, data, index)
                badge = page.locator("div.border-rd-full.px-3.py-1").first
                badge.wait_for(state="visible", timeout=15000)
                print(badge.inner_text().strip())
                if badge.inner_text().strip() == "Dewasa":
                    gender_locator = (
                        page.locator("div.flex.flex-col.gap-2")
                        .filter(has_text="Jenis Kelamin")
                        .locator("div.font-bold")
                    )
                    gender = gender_locator.inner_text().strip()
                    if gender == "Laki-Laki":
                        print("Skrining Laki-Laki Dewasa")
                        print("============== Skrining Mandiri Dimulai ==============")
                        do_demografi_dewasa(page, data, index)
                        do_risiko_kanker_usus(page, data, index)
                        do_risiko_tb(page, data, index)
                        do_hati(page, data, index)
                        do_keswa(page, data, index)
                        do_risiko_kanker_paru(page, data, index)
                        do_perilaku_merokok(page, data, index)
                        do_aktivitas_fisik(page, data, index)
                        print("============== Skrining Mandiri Selesai ==============")
                        print("============== Skrining Oleh Nakes Dimulai ==============")
                        do_gizi_laki(page, data, index)
                        do_gula_darah_dewasa(page, data, index)
                        do_tekanan_darah_dewasa(page, data, index)
                        do_risiko_tb(page, data, index)
                        do_tb(page, data, index)
                        do_frambusia(page, data, index)
                        do_kusta(page, data, index)
                        do_skabies(page, data, index)
                        do_telinga_mata(page, data, index)
                        do_karies(page, data, index)
                        do_periodontal(page, data, index)
                        do_ppok(page, data, index)
                        do_kadar_co(page, data, index)
                        do_lipid(page, data, index)
                        do_fibrosis(page, data, index)
                        do_hepatitis(page, data, index)
                        do_fungsi_ginjal(page, data, index)
                        do_kerusakan_ginjal(page, data, index)
                        do_jantung(page, data, index)
                        do_kanker_usus(page, data, index)
                        do_kanker_paru(page, data, index)
                        do_hiv(page, data, index)
                        do_sifilis(page, data, index)
                        print("============== Skrining Oleh Nakes Selesai ==============")
                        # page.pause()
                    elif gender == "Perempuan":
                        print("Skrining Perempuan Dewasa")
                        print("============== Skrining Mandiri Dimulai ==============")
                        do_demografi_dewasa_perempuan(page, data, index)
                        do_risiko_kanker_usus(page, data, index)
                        do_risiko_tb(page, data, index)
                        do_hati(page, data, index)
                        do_leher_rahim(page, data, index)
                        do_keswa(page, data, index)
                        do_risiko_kanker_paru(page, data, index)
                        do_perilaku_merokok(page, data, index)
                        do_aktivitas_fisik(page, data, index)
                        print("============== Skrining Mandiri Selesai ==============")
                        print("============== Skrining Oleh Nakes Dimulai ==============")
                        do_gizi_perempuan(page, data, index)
                        do_gula_darah_dewasa(page, data, index)
                        do_tekanan_darah_dewasa(page, data, index)
                        do_risiko_tb(page, data, index)
                        do_tb(page, data, index)
                        do_frambusia(page, data, index)
                        do_kusta(page, data, index)
                        do_skabies(page, data, index)
                        do_telinga_mata(page, data, index)
                        do_karies(page, data, index)
                        do_periodontal(page, data, index)
                        do_ppok(page, data, index)
                        do_kadar_co(page, data, index)
                        do_lipid(page, data, index)
                        do_fibrosis(page, data, index)
                        do_hepatitis(page, data, index)
                        do_fungsi_ginjal_perempuan(page, data, index)
                        do_kerusakan_ginjal(page, data, index)
                        do_kanker_payudara(page, data, index)
                        do_hpv_dna(page, data, index)
                        do_inspekulo_iva(page, data, index)
                        do_jantung(page, data, index)
                        do_kanker_usus(page, data, index)
                        do_kanker_paru(page, data, index)
                        do_catin_perempuan(page, data, index)
                        do_hiv(page, data, index)
                        do_sifilis(page, data, index)
                        print("============== Skrining Oleh Nakes Selesai ==============")
                elif badge.inner_text().strip() == "Lansia":
                    gender_locator = (
                        page.locator("div.flex.flex-col.gap-2")
                        .filter(has_text="Jenis Kelamin")
                        .locator("div.font-bold")
                    )
                    gender = gender_locator.inner_text().strip()
                    if gender == "Laki-Laki":
                        print("Skrining Lansia Laki-Laki")
                    elif gender == "Perempuan":
                        print("Skrining Lansia Perempuan")

                page.wait_for_load_state("networkidle")

                update_row_status(
                    workbook, sheet, headers, excel_path, index, "SUCCESS"
                )
            except Exception as exc:
                failed_rows.append(index)
                traceback.print_exc()
                update_row_status(
                    workbook, sheet, headers, excel_path, index, f"FAILED: {exc}"
                )
                if os.getenv(DEBUG_RAISE_ERRORS_ENV) == "1":
                    raise

        context.close()
        browser.close()


if __name__ == "__main__":
    main()
