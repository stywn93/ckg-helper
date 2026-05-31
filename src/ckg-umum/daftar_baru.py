import os

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright_window_layout import launch_chromium_with_layout, pause_with_inspector_layout

load_dotenv()

STATUS_COLUMN = "status"
USERNAME_ENV = "CKG_USERNAME"
PASSWORD_ENV = "CKG_PASSWORD"
MONTH_LABELS = {
    "01": "Jan",
    "02": "Feb",
    "03": "Mar",
    "04": "Apr",
    "05": "Mei",
    "06": "Jun",
    "07": "Jul",
    "08": "Agt",
    "09": "Sep",
    "10": "Okt",
    "11": "Nov",
    "12": "Des",
}
MONTH_TO_NUMBER = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "Mei": 5,
    "Jun": 6,
    "Jul": 7,
    "Agt": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Des": 12,
}

def is_success_status(value) -> bool:
    return str(value).strip().upper() == "SUCCESS"


def get_required_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Environment variable {name} belum diisi.")
    return value


def load_rows_from_excel(path: str) -> tuple:
    workbook = load_workbook(path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if STATUS_COLUMN not in headers:
        sheet.cell(row=1, column=len(headers) + 1, value=STATUS_COLUMN)
        headers.append(STATUS_COLUMN)

    rows = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        row_data = dict(zip(headers, row))
        if is_success_status(row_data.get(STATUS_COLUMN)):
            continue
        rows.append({"row_number": row_number, "data": row_data})

    return workbook, sheet, headers, rows


def select_date_from_picker(page, field_selector: str, date_value: str) -> None:
    year, month, day = date_value.split("-")
    target_year = int(year)
    target_month = int(month)
    month_label = MONTH_LABELS[month]

    page.locator(field_selector).click()

    popup = page.locator(".mx-datepicker-popup")
    month_btn = popup.locator(".mx-btn-current-month")
    prev_month = popup.locator(".mx-btn-icon-left")
    next_month = popup.locator(".mx-btn-icon-right")
    year_btn = popup.locator(".mx-btn-current-year")
    prev_year = popup.locator(".mx-btn-icon-double-left")

    while int(year_btn.text_content().strip()) != year:
        current_year = int(year_btn.text_content().strip())
        if target_year < current_year:
            prev_year.click()
        else:
            break


    while True:
        current_label = month_btn.text_content().strip()
        if current_label == month_label:
            break
        current_month = int(MONTH_TO_NUMBER[current_label])
        if target_month < current_month:
            prev_month.click()
        else:
            next_month.click()

        page.wait_for_timeout(300)

    popup.locator(f'td.cell[title="{date_value}"]').click()

    # logic
    # ambil dulu tahun sekarang, ambil tahun lahir
    # bandingkan tahun lahir dengan tahun sekarang, pasti minimal sama atau lebih kecil
    # selama belum sama, maka ulangi
    # year_btn.text_content().strip() != year
    # setelah tahun cocok, lanjut ke bulan
    # ambil dulu bulan sekarang, konversi ke angka, mei = 5
    # ambil bulan lahir
    # bandingkan antara bulan sekarang dengan bulan lahir pasien
    # jika bulan lahir pasien lebih besar daripada bulan sekarang, maka klik mx-btn-icon-right
    # ulangi sampai bulan lahir pasien sama dengan bulan sekarang
    # year_btn.text_content().strip() != year
    # jika demikian berarti ada while dalam while bukan?

def select_date_from_picker2(trigger_locator, date_value: str) -> None:
    year, month, day = date_value.split("-")
    target_year = int(year)
    target_month = int(month)
    month_label = MONTH_LABELS[month]

    trigger_locator.click()

    popup = trigger_locator.page.locator(".mx-datepicker-popup")
    month_btn = popup.locator(".mx-btn-current-month")
    prev_month = popup.locator(".mx-btn-icon-left")
    next_month = popup.locator(".mx-btn-icon-right")
    year_btn = popup.locator(".mx-btn-current-year")
    prev_year = popup.locator(".mx-btn-icon-double-left")
    # print("just before entering while condition")
    # trigger_locator.page.pause()
    while int(year_btn.text_content().strip()) != year:
        current_year = int(year_btn.text_content().strip())
        if target_year < current_year:
            prev_year.click()
        else:
            break

    # print("entering month while condition")
    # trigger_locator.page.pause()
    while True:
        current_label = month_btn.text_content().strip()
        if current_label == month_label:
            break
        current_month = int(MONTH_TO_NUMBER[current_label])
        if target_month < current_month:
            prev_month.click()
        else:
            next_month.click()

        trigger_locator.page.wait_for_timeout(300)

    popup.locator(f'td.cell[title="{date_value}"]').click()

def format_cell_value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def update_row_status(workbook, sheet, headers: list, excel_path: str, row_number: int, status: str) -> None:
    status_column_index = headers.index(STATUS_COLUMN) + 1
    sheet.cell(row=row_number, column=status_column_index, value=status)
    workbook.save(excel_path)


def prepare_registration_page(page) -> None:
    page.goto("https://sehatindonesiaku.kemkes.go.id/ckg-pendaftaran-individu")
    page.wait_for_load_state("networkidle")
    # page.reload(wait_until="networkidle")

    checkbox = page.locator("input[name='verify']")
    if checkbox.count() > 0:
        checkbox = page.locator("input[name='verify']")
        checkbox.set_checked(True, force=True)
        page.locator("button:has-text('Setuju')").click()
        page.wait_for_load_state("networkidle")

    page.get_by_role("button", name="Daftar Baru").click()


def register_single_entry(page, data: dict, row_number: int) -> None:
    prepare_registration_page(page)

    nik_input = page.locator("form input#nik")
    nik_input.fill(format_cell_value(data["nik"]))
    page.locator('input#Nama\\ Lengkap').fill(format_cell_value(data["nama_lengkap"]))

    select_date_from_picker(page, "#Tanggal\\ Lahir .mx-input-wrapper", format_cell_value(data["tgl_lahir"]))
    page.get_by_text("Pilih jenis kelamin", exact=True).click()
    page.locator("div.absolute.top-13.z-2000").get_by_text(
        format_cell_value(data["gender"]),
        exact=True,
    ).click()
    
    page.locator('input#No\\ Whatsapp').fill(format_cell_value(data["no_whatsapp"]))

    panel = page.locator("div:has(> .text-\\[20px\\].font-bold:text('Tanggal Pemeriksaan'))")
    panel.get_by_role("button", name=format_cell_value(data["tgl_pemeriksaan"])).nth(1).click()

    page.get_by_role("button", name="Selanjutnya").click()
    btn_recheck = page.locator("button:has-text('Periksa Kembali')").first
    btn_success = page.locator("button:has-text('Lanjutkan')").first
    try:
        btn_recheck.wait_for(state="visible", timeout=3000)
        btn_recheck.click()
        btn_recheck.wait_for(state="hidden", timeout=5000)
        checkbox = page.locator("input[name='noNik']")
        checkbox.set_checked(True, force=True)
        page.locator("input#nik\\ wali").fill(format_cell_value(data["nik_wali"]))
        page.locator('input[name="Nama Lengkap Wali"]').fill(format_cell_value(data["nama_wali"]))

        page.locator('[id="Tanggal Lahir"] .mx-input-wrapper').filter(has_text="Pilih Tanggal Lahir").click()
        print("tgl lahir wali picker clicked")
        # page.pause()
        select_date_from_picker2(
            page.locator('[id="Tanggal Lahir"] .mx-input-wrapper').filter(has_text="Pilih Tanggal Lahir"),
            format_cell_value(data["tgl_lahir_wali"]),
        )

        page.locator("div:has(> .text-gray-4:text('Pilih Jenis Kelamin'))").click()
        page.locator(".max-h-\\[250px\\]").get_by_text(format_cell_value(data["gender_wali"]), exact=True).click()
        page.locator("label").filter(has_text="No. Whatsapp Wali").locator('input[name="Nomor whatsapp"]').fill(
            format_cell_value(data["no_whatsapp_wali"])
        )
        page.get_by_role("button", name="Selanjutnya").click()
        page.locator("button:has-text('Lanjutkan')").click()

    except PlaywrightTimeoutError:
        btn_success.click()

    page.get_by_text("Pilih status pernikahan", exact=True).click()
    page.get_by_text(format_cell_value(data["pernikahan"]), exact=True).click()

    page.get_by_text("Pilih pekerjaan", exact=True).click()
    page.get_by_text(format_cell_value(data["pekerjaan"]), exact=True).click()

    page.get_by_text("Pilih alamat domisili", exact=True).click()
    page.get_by_text(format_cell_value(data["prov"]), exact=True).click()
    page.get_by_text(format_cell_value(data["kab"]), exact=True).click()
    page.get_by_text(format_cell_value(data["kec"]), exact=True).click()
    page.get_by_text(format_cell_value(data["desa"]), exact=True).click()
    page.locator("textarea#detail-domisili").fill(format_cell_value(data["domisili"]))

    page.get_by_role("button", name="Selanjutnya").click()
    page.wait_for_timeout(1500)
    page.get_by_role("button", name="Daftarkan Tanpa NIK").click()
    page.wait_for_timeout(1500)
    page.wait_for_load_state("networkidle")
    page.get_by_role("button", name="Tutup").click()
    # page.goto("https://sehatindonesiaku.kemkes.go.id/ckg-pelayanan")
    # page.wait_for_load_state("networkidle")

def main():
    excel_path = "dataset/pendaftaran_umum.xlsx"
    username = get_required_env(USERNAME_ENV)
    password = get_required_env(PASSWORD_ENV)
    workbook, sheet, headers, data_rows = load_rows_from_excel(excel_path)
    if not data_rows:
        print("Tidak ada data pada file Excel.")
        return

    with sync_playwright() as p:
        browser, window_layout = launch_chromium_with_layout(p)
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        page.goto("https://sehatindonesiaku.kemkes.go.id/login")
        page.locator("input#email").fill(username)
        page.locator("input#password").fill(password)
        pause_with_inspector_layout(page, window_layout)

        failed_rows = []

        for row_entry in data_rows:
            index = row_entry["row_number"]
            data = row_entry["data"]
            try:
                register_single_entry(page, data, index)
                update_row_status(workbook, sheet, headers, excel_path, index, "SUCCESS")
            except Exception as exc:
                failed_rows.append(index)
                update_row_status(workbook, sheet, headers, excel_path, index, f"FAILED: {exc}")
                # print(f"Baris Excel {index} gagal diproses: {exc}")

        context.close()
        browser.close()

if __name__ == "__main__":
    main()
