import os

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright_window_layout import launch_chromium_with_layout, pause_with_inspector_layout

load_dotenv()

STATUS_COLUMN = "status"
USERNAME_ENV = "CKG_USERNAME"
PASSWORD_ENV = "CKG_PASSWORD"
MONTH_LABELS = {
    "01": "Jan",
    "02": "Feb",
    "03": "Mar",
    "04": "Apr",
    "05": "Mei",
    "06": "Jun",
    "07": "Jul",
    "08": "Agt",
    "09": "Sep",
    "10": "Okt",
    "11": "Nov",
    "12": "Des",
}
MONTH_TO_NUMBER = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "Mei": 5,
    "Jun": 6,
    "Jul": 7,
    "Agt": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Des": 12,
}

def is_success_status(value) -> bool:
    return str(value).strip().upper() == "SUCCESS"


def get_required_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Environment variable {name} belum diisi.")
    return value


def load_rows_from_excel(path: str) -> tuple:
    workbook = load_workbook(path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if STATUS_COLUMN not in headers:
        sheet.cell(row=1, column=len(headers) + 1, value=STATUS_COLUMN)
        headers.append(STATUS_COLUMN)

    rows = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        row_data = dict(zip(headers, row))
        if is_success_status(row_data.get(STATUS_COLUMN)):
            continue
        rows.append({"row_number": row_number, "data": row_data})

    return workbook, sheet, headers, rows


def select_date_from_picker(page, field_selector: str, date_value: str) -> None:
    year, month, day = date_value.split("-")
    target_year = int(year)
    target_month = int(month)
    month_label = MONTH_LABELS[month]

    page.locator(field_selector).click()

    popup = page.locator(".mx-datepicker-popup")
    month_btn = popup.locator(".mx-btn-current-month")
    prev_month = popup.locator(".mx-btn-icon-left")
    next_month = popup.locator(".mx-btn-icon-right")
    year_btn = popup.locator(".mx-btn-current-year")
    prev_year = popup.locator(".mx-btn-icon-double-left")

    while int(year_btn.text_content().strip()) != year:
        current_year = int(year_btn.text_content().strip())
        if target_year < current_year:
            prev_year.click()
        else:
            break


    while True:
        current_label = month_btn.text_content().strip()
        if current_label == month_label:
            break
        current_month = int(MONTH_TO_NUMBER[current_label])
        if target_month < current_month:
            prev_month.click()
        else:
            next_month.click()

        page.wait_for_timeout(300)

    popup.locator(f'td.cell[title="{date_value}"]').click()

def set_search_date(trigger_locator, date_value: str) -> None:
    year, month, day = date_value.split("-")
    target_year = int(year)
    target_month = int(month)
    month_label = MONTH_LABELS[month]

    trigger_locator.click()

    popup = trigger_locator.page.locator(".mx-datepicker-popup")
    month_btn = popup.locator(".mx-btn-current-month")
    prev_month = popup.locator(".mx-btn-icon-left")
    next_month = popup.locator(".mx-btn-icon-right")
    year_btn = popup.locator(".mx-btn-current-year")
    prev_year = popup.locator(".mx-btn-icon-double-left")

    while int(year_btn.text_content().strip()) != year:
        current_year = int(year_btn.text_content().strip())
        if target_year < current_year:
            prev_year.click()
        else:
            break

    while True:
        current_label = month_btn.text_content().strip()
        if current_label == month_label:
            break
        current_month = int(MONTH_TO_NUMBER[current_label])
        if target_month < current_month:
            prev_month.click()
        else:
            next_month.click()

        trigger_locator.page.wait_for_timeout(300)

    popup.locator(f'td.cell[title="{date_value}"]').click()


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def update_row_status(workbook, sheet, headers: list, excel_path: str, row_number: int, status: str) -> None:
    status_column_index = headers.index(STATUS_COLUMN) + 1
    sheet.cell(row=row_number, column=status_column_index, value=status)
    workbook.save(excel_path)


def prepare_registration_page(page) -> None:
    page.goto("https://sehatindonesiaku.kemkes.go.id/ckg-pendaftaran-individu")
    page.wait_for_load_state("networkidle")

    checkbox = page.locator("input[name='verify']")
    if checkbox.count() > 0:
        checkbox = page.locator("input[name='verify']")
        checkbox.set_checked(True, force=True)
        page.locator("button:has-text('Setuju')").click()
        page.wait_for_load_state("networkidle")


def searchPatient(page, data: dict, row_number: int, window_layout) -> None:
    prepare_registration_page(page)

    set_search_date(page.locator('[id="Tanggal Pemeriksaan"]'), format_cell_value(data["tgl_pemeriksaan"]))

    page.locator("span:has-text('Nomor Tiket')").click()
    page.get_by_text("Nama", exact=True).click()
    page.locator('input#searchNik').fill(format_cell_value(data["nama_lengkap"]))
    page.keyboard.press("Enter")
    page.wait_for_load_state("networkidle")
    confirm_button = page.locator("button:has-text('Konfirmasi Hadir')").first
    try:
        confirm_button.wait_for(state="visible", timeout=3000)
    except PlaywrightTimeoutError:
        raise RuntimeError("Pencarian tidak memberikan hasil atau tombol Konfirmasi Hadir tidak muncul.")

    confirm_button.click()
    checkbox = page.locator("input[name='verify']")
    if checkbox.count() > 0:
        checkbox = page.locator("input[name='verify']")
        checkbox.set_checked(True, force=True)
        hadir_button = page.get_by_role("button", name="Hadir", exact=True)
        hadir_button.wait_for(state="visible", timeout=3000)
        hadir_button.click()
        page.wait_for_load_state("networkidle")
        page.get_by_role("button", name="Tutup", exact=True).click()

    #inspect again
    pause_with_inspector_layout(page, window_layout)



def main():
    excel_path = "dataset/konfirm_kehadiran.xlsx"
    username = get_required_env(USERNAME_ENV)
    password = get_required_env(PASSWORD_ENV)
    workbook, sheet, headers, data_rows = load_rows_from_excel(excel_path)
    if not data_rows:
        print("Tidak ada data pada file Excel.")
        return

    with sync_playwright() as p:
        browser, window_layout = launch_chromium_with_layout(p)
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        page.goto("https://sehatindonesiaku.kemkes.go.id/login")
        page.locator("input#email").fill(username)
        page.locator("input#password").fill(password)
        pause_with_inspector_layout(page, window_layout)

        failed_rows = []

        for row_entry in data_rows:
            index = row_entry["row_number"]
            data = row_entry["data"]
            try:
                searchPatient(page, data, index, window_layout)
                update_row_status(workbook, sheet, headers, excel_path, index, "SUCCESS")
            except Exception as exc:
                failed_rows.append(index)
                update_row_status(workbook, sheet, headers, excel_path, index, f"FAILED: {exc}")
                # print(f"Baris Excel {index} gagal diproses: {exc}")

        context.close()
        browser.close()

if __name__ == "__main__":
    main()
