from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

STATUS_COLUMN = "status"
TRACE_DIR = Path("traces")
MONTH_LABELS = {
    "01": "Jan",
    "02": "Feb",
    "03": "Mar",
    "04": "Apr",
    "05": "Mei",
    "06": "Jun",
    "07": "Jul",
    "08": "Agu",
    "09": "Sep",
    "10": "Okt",
    "11": "Nov",
    "12": "Des",
}


def load_rows_from_excel(path: str) -> tuple:
    workbook = load_workbook(path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if STATUS_COLUMN not in headers:
        sheet.cell(row=1, column=len(headers) + 1, value=STATUS_COLUMN)
        headers.append(STATUS_COLUMN)

    rows = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        row_data = dict(zip(headers, row))
        rows.append({"row_number": row_number, "data": row_data})

    return workbook, sheet, headers, rows


def select_date_from_picker(page, field_selector: str, date_value: str) -> None:
    year, month, day = date_value.split("-")
    month_label = MONTH_LABELS[month]

    page.locator(field_selector).click()

    popup = page.locator(".mx-datepicker-popup")
    month_btn = popup.locator(".mx-btn-current-month")
    prev_month = popup.locator(".mx-btn-icon-left")
    year_btn = popup.locator(".mx-btn-current-year")
    prev_year = popup.locator(".mx-btn-icon-double-left")

    while month_btn.text_content().strip() != month_label:
        prev_month.click()

    while year_btn.text_content().strip() != year:
        prev_year.click()

    popup.locator(f'td.cell[title="{date_value}"]').click()


def select_date_from_picker2(trigger_locator, date_value: str) -> None:
    year, month, day = date_value.split("-")
    month_label = MONTH_LABELS[month]

    trigger_locator.click()

    popup = trigger_locator.page.locator(".mx-datepicker-popup")
    month_btn = popup.locator(".mx-btn-current-month")
    prev_month = popup.locator(".mx-btn-icon-left")
    year_btn = popup.locator(".mx-btn-current-year")
    prev_year = popup.locator(".mx-btn-icon-double-left")

    while month_btn.text_content().strip() != month_label:
        prev_month.click()

    while year_btn.text_content().strip() != year:
        prev_year.click()

    popup.locator(f'td.cell[title="{date_value}"]').click()


def format_cell_value(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)


def update_row_status(workbook, sheet, headers: list, excel_path: str, row_number: int, status: str) -> None:
    status_column_index = headers.index(STATUS_COLUMN) + 1
    sheet.cell(row=row_number, column=status_column_index, value=status)
    workbook.save(excel_path)


def build_trace_path() -> Path:
    TRACE_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return TRACE_DIR / f"automation-trace-{timestamp}.zip"


def prepare_registration_page(page) -> None:
    page.goto("https://sehatindonesiaku.kemkes.go.id/ckg-pendaftaran-individu")
    page.wait_for_load_state("networkidle")
    page.reload(wait_until="networkidle")

    checkbox = page.locator("input[name='verify']")
    if checkbox.count() > 0:
        checkbox = page.locator("input[name='verify']")
        checkbox.set_checked(True, force=True)
        page.locator("button:has-text('Setuju')").click()
        page.wait_for_load_state("networkidle")

    page.get_by_role("button", name="Daftar Baru").click()


def register_single_entry(page, data: dict, row_number: int) -> None:
    print(f"Memproses baris Excel {row_number}...")
    prepare_registration_page(page)

    nik_input = page.locator("form input#nik")
    nik_input.fill(format_cell_value(data["nik"]))
    page.locator('input#Nama\\ Lengkap').fill(format_cell_value(data["nama_lengkap"]))

    #masih menyisakan PR jika bulannya lebih kecil dari bulan ini
    select_date_from_picker(page, "#Tanggal\\ Lahir .mx-input-wrapper", format_cell_value(data["tgl_lahir"]))

    # page.get_by_text("Pilih jenis kelamin", exact=True).click()
    # page.get_by_text(format_cell_value(data["gender"]), exact=True).click()
    page.get_by_text("Pilih jenis kelamin", exact=True).click()
    page.locator("div.absolute.top-13.z-2000").get_by_text(
        format_cell_value(data["gender"]),
        exact=True,
    ).click()
    
    page.locator('input#No\\ Whatsapp').fill(format_cell_value(data["no_whatsapp"]))

    panel = page.locator("div:has(> .text-\\[20px\\].font-bold:text('Tanggal Pemeriksaan'))")
    panel.get_by_role("button", name=format_cell_value(data["tgl_pemeriksaan"])).nth(1).click()

    page.get_by_role("button", name="Selanjutnya").click()
    btn_recheck = page.locator("button:has-text('Periksa Kembali')").first
    btn_success = page.locator("button:has-text('Lanjutkan')").first
    try:
        btn_recheck.wait_for(state="visible", timeout=3000)
        btn_recheck.click()
        btn_recheck.wait_for(state="hidden", timeout=5000)
        checkbox = page.locator("input[name='noNik']")
        checkbox.set_checked(True, force=True)
        page.locator("input#nik\\ wali").fill(format_cell_value(data["nik_wali"]))
        page.locator('input[name="Nama Lengkap Wali"]').fill(format_cell_value(data["nama_wali"]))

        page.locator('[id="Tanggal Lahir"] .mx-input-wrapper').filter(has_text="Pilih Tanggal Lahir").click()

        select_date_from_picker2(
            page.locator('[id="Tanggal Lahir"] .mx-input-wrapper').filter(has_text="Pilih Tanggal Lahir"),
            format_cell_value(data["tgl_lahir_wali"]),
        )

        page.locator("div:has(> .text-gray-4:text('Pilih Jenis Kelamin'))").click()
        page.locator(".max-h-\\[250px\\]").get_by_text(format_cell_value(data["gender_wali"]), exact=True).click()
        page.locator("label").filter(has_text="No. Whatsapp Wali").locator('input[name="Nomor whatsapp"]').fill(
            format_cell_value(data["no_whatsapp_wali"])
        )
        page.get_by_role("button", name="Selanjutnya").click()
        page.locator("button:has-text('Lanjutkan')").click()

    except PlaywrightTimeoutError:
        btn_success.click()

    page.get_by_text("Pilih status pernikahan", exact=True).click()
    page.get_by_text(format_cell_value(data["pernikahan"]), exact=True).click()

    page.get_by_text("Pilih pekerjaan", exact=True).click()
    page.get_by_text(format_cell_value(data["pekerjaan"]), exact=True).click()

    page.get_by_text("Pilih alamat domisili", exact=True).click()
    page.get_by_text(format_cell_value(data["prov"]), exact=True).click()
    page.get_by_text(format_cell_value(data["kab"]), exact=True).click()
    page.get_by_text(format_cell_value(data["kec"]), exact=True).click()
    page.get_by_text(format_cell_value(data["desa"]), exact=True).click()

    page.locator("textarea#detail-domisili").fill(format_cell_value(data["domisili"]))

    page.get_by_role("button", name="Selanjutnya").click()
    page.wait_for_timeout(1500)
    page.get_by_role("button", name="Daftarkan Tanpa NIK").click()
    page.wait_for_timeout(1500)
    page.wait_for_load_state("networkidle")
    page.get_by_role("button", name="Tutup").click()
    print(f"Baris Excel {row_number} berhasil didaftarkan.")


def main():
    excel_path = "data.xlsx"
    workbook, sheet, headers, data_rows = load_rows_from_excel(excel_path)
    if not data_rows:
        print("Tidak ada data pada file Excel.")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(no_viewport=True)
        page = context.new_page()
        trace_path = build_trace_path()

        page.goto("https://sehatindonesiaku.kemkes.go.id/login")
        page.locator("input#email").fill("asembagusjempol@gmail.com")
        page.locator("input#password").fill("Asembagus*1")
        page.pause()
        context.tracing.start(screenshots=True, snapshots=True, sources=True)

        failed_rows = []

        try:
            for row_entry in data_rows:
                index = row_entry["row_number"]
                data = row_entry["data"]
                try:
                    register_single_entry(page, data, index)
                    update_row_status(workbook, sheet, headers, excel_path, index, "SUCCESS")
                except Exception as exc:
                    failed_rows.append(index)
                    update_row_status(workbook, sheet, headers, excel_path, index, f"FAILED: {exc}")
                    print(f"Baris Excel {index} gagal diproses: {exc}")

            if failed_rows:
                print(f"Selesai dengan kegagalan pada baris: {failed_rows}")
            else:
                print("Semua baris Excel berhasil diproses.")
        finally:
            context.tracing.stop(path=str(trace_path))
            print(f"Playwright trace disimpan di: {trace_path}")
            context.close()
            browser.close()

if __name__ == "__main__":
    main()
